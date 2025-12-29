#!/usr/bin/env python3
"""Pre-render `stock-details.html` into a crawlable static HTML file.

This script reads the same XLSX dataset your front-end uses
(`data/financials_analysis_dashboard_offset_0.xlsx`), computes medians +
section scores the same way as `js/stock_details_page.js`, and writes an
HTML file with all metrics already filled (no JS required).

Usage:
  python prerender_stock_details.py --symbol 'AAPL:US'

Output:
    HTML_portfolio/stocks/AAPL-US.html
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from bs4 import BeautifulSoup
from openpyxl import load_workbook


DATA_FIELDS: list[str] = [
    "tang_equity_over_tot_liab",
    "capital_intensity_reverse",
    "cagr_tangible_book_per_share",
    "cagr_cash_and_equiv",
    "goodwill_to_assets",
    "leverage_ratio",
    "interest_coverage_ratio",
    "roe_tangible_equity",
    "roic_over_wacc",
    "rule_of_40",
    "cash_conversion_ratio",
    "earnings_yield",
    "price_to_earnings",
    "fcf_yield",
    "peg",
    "price_to_tangible_book",
    "relative_PE_vs_history",
    "avg_5years_eps_growth",
    "avg_5years_revenue_growth",
    "revenue_growth_acceleration",
    "cagr_shares_diluted",
    "expected_growth_market_cap_10Y",
    "final_earnings_for_10y_growth_10perc",
    "final_earnings_for_10y_growth_15perc",
    "avg_5years_roe_growth",
    "implied_perpetual_growth_curr_market_cap",
    "current_ratio",
    "negative_eps_count_5y",
    "eps_growth_5y_total",
    "pe_times_pb",
]

SECTION_FIELDS: dict[str, list[str]] = {
    "balance-sheet-strength": [
        "tang_equity_over_tot_liab",
        "capital_intensity_reverse",
        "cagr_tangible_book_per_share",
        "cagr_cash_and_equiv",
        "goodwill_to_assets",
    ],
    "debt-service": ["leverage_ratio", "interest_coverage_ratio"],
    "profitability": [
        "roe_tangible_equity",
        "roic_over_wacc",
        "rule_of_40",
        "cash_conversion_ratio",
        "avg_5years_roe_growth",
    ],
    "valuation": [
        "earnings_yield",
        "price_to_earnings",
        "fcf_yield",
        "peg",
        "price_to_tangible_book",
        "relative_PE_vs_history",
    ],
    "growth": [
        "avg_5years_eps_growth",
        "avg_5years_revenue_growth",
        "revenue_growth_acceleration",
        "cagr_shares_diluted",
        "expected_growth_market_cap_10Y",
        "final_earnings_for_10y_growth_10perc",
        "final_earnings_for_10y_growth_15perc",
        "implied_perpetual_growth_curr_market_cap",
    ],
    "graham-value-investor-indicator": [
        "current_ratio",
        "negative_eps_count_5y",
        "eps_growth_5y_total",
        "pe_times_pb",
    ],
}

HIGHER_IS_BETTER: set[str] = {
    "tang_equity_over_tot_liab",
    "capital_intensity_reverse",
    "cagr_tangible_book_per_share",
    "cagr_cash_and_equiv",
    "roe_tangible_equity",
    "roic_over_wacc",
    "rule_of_40",
    "cash_conversion_ratio",
    "earnings_yield",
    "fcf_yield",
    "avg_5years_eps_growth",
    "avg_5years_revenue_growth",
    "revenue_growth_acceleration",
    "expected_growth_market_cap_10Y",
    "avg_5years_roe_growth",
    "interest_coverage_ratio",
    "current_ratio",
}

LOWER_IS_BETTER: set[str] = {
    "price_to_earnings",
    "peg",
    "price_to_tangible_book",
    "final_earnings_for_10y_growth_10perc",
    "final_earnings_for_10y_growth_15perc",
    "implied_perpetual_growth_curr_market_cap",
    "leverage_ratio",
    "goodwill_to_assets",
    "relative_PE_vs_history",
    "cagr_shares_diluted",
    "negative_eps_count_5y",
}

MEDIAN_MARGIN_OF_SAFETY = 0.15


def sanitize_symbol_for_filename(symbol: str) -> str:
    """Make a symbol safe for filenames/URLs.

    Examples:
        AAPL:US -> AAPL-US
        BRK.B:US -> BRK.B-US (keeps dot)
    """

    s = symbol.strip().upper()
    # Replace common separators that can be problematic in paths/URLs
    s = s.replace(":", "-")
    s = s.replace("/", "-")
    s = s.replace("\\", "-")
    # Collapse whitespace to '-'
    s = "-".join(s.split())
    # Remove any characters that are still risky
    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-._")
    s = "".join(ch for ch in s if ch in allowed)
    # Avoid empty output
    return s or "SYMBOL"


@dataclass(frozen=True)
class Indicator:
    icon: str
    css_class: str


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s or s.lower() in {"na", "n/a", "nan", "null"}:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def calculate_median(values: Iterable[float | None]) -> float | None:
    clean = [
        v
        for v in values
        if isinstance(v, (int, float)) and v is not None and not math.isnan(float(v))
    ]
    clean.sort()
    if not clean:
        return None
    mid = len(clean) // 2
    if len(clean) % 2 == 0:
        return (clean[mid - 1] + clean[mid]) / 2.0
    return clean[mid]


def format_metric_value(value: float | None, field_id: str) -> str:
    if value is None or not isinstance(value, (int, float)) or math.isnan(float(value)):
        return "N/A"

    if field_id in {
        "tang_equity_over_tot_liab",
        "capital_intensity_reverse",
        "roic_over_wacc",
        "price_to_earnings",
        "peg",
        "price_to_tangible_book",
        "leverage_ratio",
        "interest_coverage_ratio",
        "relative_PE_vs_history",
        "current_ratio",
        "pe_times_pb",
    }:
        return f"{value:.2f}"

    if field_id in {"negative_eps_count_5y"}:
        return f"{value:.0f}"

    if field_id in {
        "cagr_tangible_book_per_share",
        "cagr_cash_and_equiv",
        "roe_tangible_equity",
        "cash_conversion_ratio",
        "earnings_yield",
        "fcf_yield",
        "avg_5years_eps_growth",
        "avg_5years_revenue_growth",
        "revenue_growth_acceleration",
        "expected_growth_market_cap_10Y",
        "avg_5years_roe_growth",
        "implied_perpetual_growth_curr_market_cap",
        "goodwill_to_assets",
        "cagr_shares_diluted",
        "eps_growth_5y_total",
    }:
        return f"{value * 100:.2f}%"

    if field_id in {
        "final_earnings_for_10y_growth_10perc",
        "final_earnings_for_10y_growth_15perc",
    }:
        return f"{value / 1_000_000_000:.2f}B"

    if field_id in {"rule_of_40"}:
        return f"{value * 100:.2f}"

    # Default: locale formatting-ish
    return f"{value:,.2f}"


def threshold_indicator(
    field_id: str,
) -> Callable[[float | None], Indicator | None] | None:
    def _rule_current_ratio(v: float | None) -> Indicator | None:
        if v is None or math.isnan(float(v)):
            return None
        return Indicator("▲", "better") if v >= 1.5 else Indicator("▼", "worse")

    def _rule_negative_eps(v: float | None) -> Indicator | None:
        if v is None or math.isnan(float(v)):
            return None
        return Indicator("▲", "better") if v == 0 else Indicator("▼", "worse")

    def _rule_eps_growth(v: float | None) -> Indicator | None:
        if v is None or math.isnan(float(v)):
            return None
        return Indicator("▲", "better") if v > 1 else Indicator("▼", "worse")

    def _rule_pe_times_pb(v: float | None) -> Indicator | None:
        if v is None or math.isnan(float(v)):
            return None
        return Indicator("▲", "better") if v < 30 else Indicator("▼", "worse")

    rules: dict[str, Callable[[float | None], Indicator | None]] = {
        "current_ratio": _rule_current_ratio,
        "negative_eps_count_5y": _rule_negative_eps,
        "eps_growth_5y_total": _rule_eps_growth,
        "pe_times_pb": _rule_pe_times_pb,
    }
    return rules.get(field_id)


def get_comparison_indicator(
    value: float | None,
    median_value: float | None,
    field_id: str,
) -> Indicator:
    # Threshold metrics override median comparison
    rule = threshold_indicator(field_id)
    if rule is not None:
        res = rule(value)
        return res if res is not None else Indicator("▬", "equal")

    if value is None or median_value is None:
        return Indicator("▬", "equal")
    if math.isnan(float(value)) or math.isnan(float(median_value)):
        return Indicator("▬", "equal")

    eps = 1e-12
    if abs(median_value) < eps:
        if field_id in HIGHER_IS_BETTER:
            if value > median_value:
                return Indicator("▲", "better")
            if value < median_value:
                return Indicator("▼", "worse")
            return Indicator("▬", "equal")
        if field_id in LOWER_IS_BETTER:
            if value < median_value:
                return Indicator("▲", "better")
            if value > median_value:
                return Indicator("▼", "worse")
            return Indicator("▬", "equal")
        return Indicator("▬", "equal")

    upper = median_value * (1 + MEDIAN_MARGIN_OF_SAFETY)
    lower = median_value * (1 - MEDIAN_MARGIN_OF_SAFETY)

    if field_id in HIGHER_IS_BETTER:
        if value >= upper:
            return Indicator("▲", "better")
        if value <= lower:
            return Indicator("▼", "worse")
        return Indicator("▬", "equal")

    if field_id in LOWER_IS_BETTER:
        if value <= lower:
            return Indicator("▲", "better")
        if value >= upper:
            return Indicator("▼", "worse")
        return Indicator("▬", "equal")

    return Indicator("▬", "equal")


def calculate_section_score(
    section_field_ids: list[str],
    stock_data: dict[str, float | None],
    medians: dict[str, float | None],
) -> tuple[int, int]:
    above = 0
    total = 0

    for field_id in section_field_ids:
        value = stock_data.get(field_id)
        median_value = medians.get(field_id)

        # threshold metrics: don't require a median
        if threshold_indicator(field_id) is not None:
            if value is None:
                continue
            total += 1
            ind = get_comparison_indicator(value, median_value, field_id)
            if ind.css_class == "better":
                above += 1
            continue

        if value is None or median_value is None:
            continue
        total += 1
        ind = get_comparison_indicator(value, median_value, field_id)
        if ind.css_class == "better":
            above += 1

    return above, total


def get_score_chip_class(above: int, total: int) -> str:
    if total == 0:
        return "mixed"
    ratio = above / total
    if ratio >= 0.7:
        return "good"
    if ratio >= 0.4:
        return "mixed"
    return "poor"


def read_xlsx_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        if not headers:
            return []
        header_names = [str(h).strip() if h is not None else "" for h in headers]

        out: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            record: dict[str, Any] = {}
            empty_count = 0
            for idx, h in enumerate(header_names):
                if not h:
                    continue
                v = row[idx] if idx < len(row) else None
                record[h] = v
                if v is None or (isinstance(v, str) and not v.strip()):
                    empty_count += 1
            # skip completely empty rows
            if empty_count >= len(record) and record:
                continue
            out.append(record)
        return out
    finally:
        wb.close()


def build_numeric_views(rows: list[dict[str, Any]]) -> list[dict[str, float | None]]:
    numeric_rows: list[dict[str, float | None]] = []
    for r in rows:
        nr: dict[str, float | None] = {}

        # Map possible ticker column names onto a single key: "symbol"
        # Some files use "ticker" or different casing.
        for k, v in r.items():
            k_norm = str(k).strip().lower() if k is not None else ""
            if k_norm in {"symbol", "ticker", "tick", "tic"}:
                nr["symbol"] = str(v).strip() if v is not None else None
            elif k in DATA_FIELDS:
                nr[k] = _to_float(v)
            else:
                pass
        numeric_rows.append(nr)
    return numeric_rows


def compute_medians(rows: list[dict[str, float | None]]) -> dict[str, float | None]:
    medians: dict[str, float | None] = {}
    for field_id in DATA_FIELDS:
        medians[field_id] = calculate_median(r.get(field_id) for r in rows)
    return medians


def find_stock_row(
    rows: list[dict[str, float | None]], symbol: str
) -> dict[str, float | None] | None:
    sym = symbol.strip().upper()
    base_sym = sym.split(":", 1)[0].strip() if ":" in sym else sym
    for r in rows:
        rs = r.get("symbol")
        if not isinstance(rs, str):
            continue
        rs_u = rs.strip().upper()

        # exact match
        if rs_u == sym:
            return r

        # match base ticker (AAPL) against either representation
        rs_base = rs_u.split(":", 1)[0].strip() if ":" in rs_u else rs_u
        if rs_base == base_sym:
            return r
    return None


def render_html(
    template_html: str,
    symbol: str,
    stock_data: dict[str, float | None],
    medians: dict[str, float | None],
    asset_prefix: str = "",
) -> str:
    soup = BeautifulSoup(template_html, "html.parser")

    # Title + header
    if soup.title is not None:
        soup.title.string = f"Stock Details - {symbol}"

    header_span = soup.find(id="stock-symbol-header")
    if header_span is not None:
        header_span.string = symbol

    # Mark body with the symbol (useful if you keep JS around)
    if soup.body is not None:
        soup.body["data-stock-symbol"] = symbol
        soup.body["data-prerendered"] = "1"

    # Ensure assets resolve when the page is served from a subfolder (e.g. /stocks/AAPL-US.html)
    # and also when opened locally as a file.
    # Examples: asset_prefix="../" for /stocks/*.html
    for link in soup.find_all("link"):
        href = (link.get("href") or "").strip()
        if href == "styles.css":
            link["href"] = f"{asset_prefix}styles.css"

    for script in soup.find_all("script"):
        src = (script.get("src") or "").strip()
        if src.startswith("js/"):
            script["src"] = f"{asset_prefix}{src}"

    for a in soup.find_all("a"):
        href = (a.get("href") or "").strip()
        if href == "index.html":
            a["href"] = f"{asset_prefix}index.html"

    # Section chips + scorecards
    for section_id, field_ids in SECTION_FIELDS.items():
        above, total = calculate_section_score(field_ids, stock_data, medians)
        chip_class = get_score_chip_class(above, total)

        # Section header chip
        section = soup.find(id=section_id)
        if section is not None:
            h2 = section.find("h2")
            if h2 is not None:
                chip = soup.new_tag("span")
                chip["class"] = ["section-score", chip_class]
                chip.string = f"{above}/{total} above median"
                h2.append(chip)

        # Scorecard
        score_val = soup.find(id=f"score-val-{section_id}")
        score_bar = soup.find(id=f"score-bar-{section_id}")
        score_label = soup.find(id=f"score-label-{section_id}")
        card = soup.find(id=f"card-{section_id}")

        if score_val is not None:
            score_val.string = f"{above}/{total}"

        if score_bar is not None:
            pct = (above / total * 100.0) if total > 0 else 0.0
            style = score_bar.get("style", "")
            # replace any existing width
            parts = [
                p.strip()
                for p in style.split(";")
                if p.strip() and not p.strip().startswith("width:")
            ]
            parts.append(f"width: {pct:.0f}%")
            score_bar["style"] = "; ".join(parts) + ";"

        if card is not None:
            existing = card.get("class", [])
            filtered = [
                c
                for c in existing
                if c not in {"status-good", "status-mixed", "status-poor"}
            ]
            filtered.append(f"status-{chip_class}")
            card["class"] = filtered

        if score_label is not None:
            label_text = "Neutral"
            if chip_class == "good":
                label_text = "Strong"
            elif chip_class == "mixed":
                label_text = "Mixed"
            elif chip_class == "poor":
                label_text = "Weak"
            score_label.string = label_text

    # Metrics
    for field_id in DATA_FIELDS:
        value = stock_data.get(field_id)
        median_value = medians.get(field_id)
        indicator = get_comparison_indicator(value, median_value, field_id)

        value_el = soup.find(id=field_id)
        if value_el is not None:
            wrapper = soup.new_tag("div")
            wrapper["class"] = ["metric-value-wrapper"]

            ind_span = soup.new_tag("span")
            ind_span["class"] = ["metric-indicator", indicator.css_class]
            ind_span["aria-label"] = indicator.css_class
            ind_span.string = indicator.icon

            val_span = soup.new_tag("span")
            val_span["class"] = ["metric-value"]
            val_span.string = format_metric_value(value, field_id)
            if indicator.css_class == "better":
                val_span["class"].append("metric-better")
            elif indicator.css_class == "worse":
                val_span["class"].append("metric-worse")

            wrapper.append(ind_span)
            wrapper.append(val_span)
            value_el.replace_with(wrapper)

        median_el = soup.find(id=f"{field_id}_median")
        if median_el is not None:
            median_el.string = format_metric_value(median_value, field_id)

    return str(soup)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Pre-render stock-details.html for SEO"
    )
    parser.add_argument(
        "--symbol", default="AAPL", help="Ticker to render (default: AAPL)"
    )
    parser.add_argument(
        "--all-from-rankings",
        action="store_true",
        help="Generate pages for every symbol found in rank_companies JSON",
    )
    parser.add_argument(
        "--rankings-json",
        default="rank_companies/rank_companies.json",
        help="Rankings JSON path (default: rank_companies/rank_companies.json)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Optional limit for batch mode (useful for quick tests)",
    )
    parser.add_argument(
        "--no-clean",
        action="store_true",
        help="Do not delete stale HTML files in the output directory during batch mode",
    )
    parser.add_argument(
        "--template",
        default="stock-details.html",
        help="Template HTML file (default: stock-details.html)",
    )
    parser.add_argument(
        "--xlsx",
        default="data/financials_analysis_dashboard_offset_0.xlsx",
        help="XLSX data file (default: data/financials_analysis_dashboard_offset_0.xlsx)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output HTML path. If omitted, uses --out-dir/<SYMBOL>.html",
    )
    parser.add_argument(
        "--out-dir",
        default="stocks",
        help="Output directory relative to site root (default: stocks)",
    )
    args = parser.parse_args()

    # VS Code / IDE convenience: when you click “Run Python File”, no CLI args
    # are provided. In that case, default to batch generation from rankings.
    if len(sys.argv) == 1:
        args.all_from_rankings = True

    root = Path(__file__).resolve().parent
    template_path = (root / args.template).resolve()
    xlsx_path = (root / args.xlsx).resolve()

    if not template_path.exists():
        raise SystemExit(f"Template not found: {template_path}")
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    template_html = template_path.read_text(encoding="utf-8")

    # Load and precompute all data once (fast for batch mode)
    rows_raw = read_xlsx_rows(xlsx_path)
    rows = build_numeric_views(rows_raw)
    medians = compute_medians(rows)

    # Compute how many levels deep the output folder is relative to the site root.
    # This lets us write asset URLs like ../styles.css for /stocks/*.html.
    out_dir_path = (root / str(args.out_dir).strip()).resolve()
    asset_prefix = ""
    try:
        rel_parent = out_dir_path.relative_to(root)
        depth = len(rel_parent.parts)
        if depth > 0:
            asset_prefix = "../" * depth
    except ValueError:
        asset_prefix = ""

    def write_one(symbol: str) -> bool:
        sym = str(symbol).strip().upper()
        if not sym:
            return False

        safe_sym = sanitize_symbol_for_filename(sym)
        out_path = (
            Path(args.out)
            if args.out and not args.all_from_rankings
            else (out_dir_path / f"{safe_sym}.html")
        )
        out_path.parent.mkdir(parents=True, exist_ok=True)

        stock_row = find_stock_row(rows, sym)
        if stock_row is None:
            return False

        rendered = render_html(
            template_html, sym, stock_row, medians, asset_prefix=asset_prefix
        )
        out_path.write_text(rendered, encoding="utf-8")
        print(f"Wrote: {out_path}")
        return True

    if args.all_from_rankings:
        rankings_path = (root / args.rankings_json).resolve()
        if not rankings_path.exists():
            raise SystemExit(f"Rankings JSON not found: {rankings_path}")

        raw = json.loads(rankings_path.read_text(encoding="utf-8"))
        if not isinstance(raw, list):
            raise SystemExit("Rankings JSON must be a list of objects")

        symbols: list[str] = []
        for obj in raw:
            if isinstance(obj, dict) and "symbol" in obj:
                symbols.append(str(obj["symbol"]))

        # Unique, stable order
        seen: set[str] = set()
        unique_symbols: list[str] = []
        for s in symbols:
            key = str(s).strip().upper()
            if not key or key in seen:
                continue
            seen.add(key)
            unique_symbols.append(key)

        if args.limit is not None:
            unique_symbols = unique_symbols[: max(0, args.limit)]

        # In batch mode, keep the /stocks folder in sync with the JSON:
        # delete any existing HTML files that no longer correspond to a symbol.
        expected_files = {
            f"{sanitize_symbol_for_filename(sym)}.html" for sym in unique_symbols
        }
        if not args.no_clean:
            out_dir_path.mkdir(parents=True, exist_ok=True)
            deleted = 0
            kept = 0
            for p in out_dir_path.glob("*.html"):
                # Only manage files directly under the output directory.
                if p.name in expected_files:
                    kept += 1
                    continue
                try:
                    p.unlink()
                    deleted += 1
                except OSError:
                    # If a file can't be deleted, just skip it.
                    pass
            if deleted:
                print(
                    f"Cleaned: removed {deleted} stale HTML file(s) from {out_dir_path}"
                )

        print(f"Batch prerender: {len(unique_symbols)} symbols → {out_dir_path}")

        ok = 0
        missing = 0
        for sym in unique_symbols:
            if write_one(sym):
                ok += 1
            else:
                missing += 1

        print(f"Done. Generated: {ok}. Missing in XLSX: {missing}.")
        return 0

    # Single-symbol mode
    symbol = str(args.symbol).strip().upper()
    if not symbol:
        raise SystemExit("--symbol is required")

    if write_one(symbol):
        return 0
    raise SystemExit(f"Symbol not found in XLSX: {symbol}")


if __name__ == "__main__":
    raise SystemExit(main())
